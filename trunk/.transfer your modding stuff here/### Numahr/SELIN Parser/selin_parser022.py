from openpyxl import load_workbook
import sys
import string


class selin_parser:
    """
    INSTALL
    Those steps are customized for Windows - but tested only on Mac OS X
    - install Python 2.x (last version)
        https://www.python.org/downloads/release/python-278/
    - install easy_install with Python
        https://adesquared.wordpress.com/2013/07/07/setting-up-python-and-easy_install-on-windows-7/
    - install openpyxl library with easy_install
        in a DOS : easy_install openpyxl
    - put this file (selin_parser.py) and the XLXS modifiers file in the same directory
    - in a DOS shell window, launch:
        python selin_parser.py my_modifier_file.xlsx
        (note: ignore the warning about "Discarded range with reserved name")
        => in the current directory, multiple TXT files are created to be used in the mod files.
    """

    def __init__(self):
        # column 4 contains the name/code of the religions
        self._religion_name_col = 'D'
        # this range contains the list of all the religion
        self._religion_row_begin = 8
        self._religion_row_end = 280

    def parse_modifiers_in_excel(self, ws, filename):

        f = open(filename, 'w')

        for religion_row in range(self._religion_row_begin, self._religion_row_end):
            f.write("#################################################\n")
            f.write(ws.cell(column=self.col2num(self._religion_name_col), row=religion_row).value+'\n')
            f.write("#################################################\n")
            self.write_modifiers(f, ws, 'character_modifier', religion_row,
                                 self.col2num('CD'), self.col2num('DR'))
            self.write_modifiers(f, ws, 'other_modifier', religion_row,
                                 self.col2num('L'), self.col2num('CB'))

        f.close()

    def write_modifiers(self, f, ws, range_name, religion_row, begin, end):
        line_prefix = '\t\t'
        f.write(line_prefix + range_name + ' = {\n')
        for col in range(begin, end+1):
            header_value = ws.cell(column=col, row=4).value
            cell_value = ws.cell(column=col, row=religion_row).value
            if header_value is None or cell_value == 0:
                continue
            if isinstance(cell_value, float):
                f.write(line_prefix + '\t' + header_value.encode('utf-8') + " = %0.2f\n" % cell_value)
            if isinstance(cell_value, int):
                f.write(line_prefix + '\t' + header_value.encode('utf-8') + " = " + str(cell_value) + '\n')
        f.write(line_prefix + '}\n')

    def parse_mercenary_titles_in_excel(self, ws, filename):
        f = open(filename, 'w')

        f.write("# HOLY ORDERS AUTOMATICALLY GENERATED AND MAINTAINED BY THE SELIN PARSER v0.22 - Do not edit by hand!!! If a change is required, change the python code or the source Matrix !!!\n\n")
        for religion_row in range(self._religion_row_begin, self._religion_row_end):
            multiplier = ws.cell(column=self.col2num('AP'), row=religion_row).value
            if multiplier is None or multiplier == 0 or multiplier < 0:
                continue  # we skip religion with no multiplier
            rcod = ws.cell(column=self.col2num(self._religion_name_col), row=religion_row).value
            capital = ws.cell(column=self.col2num('AO'), row=religion_row).value
            color = ws.cell(column=self.col2num('EQ'), row=religion_row).value
            f.write("d_holy"+rcod+" = {\n")
            f.write('\tcolor = { '+str(color)+' }\n' +
                    '\tcolor2 = { 255 255 255 }\n' +
                    '\tcapital = '+str(capital)+'\n\n' +
                    '\tholy_order = yes\n' +
                    '\ttitle = "GRANDMASTER"\n' +
                    '\tfoa = "GRANDMASTER_FOA"\n\n' +
                    '\t# Always exists\n' +
                    '\tlandless = yes\n\n' +
                    '\t# Parent Religion/Culture\n' +
                    '\treligion = '+rcod+'\n\n' +
                    '\t# Cannot be held as a secondary title\n' +
                    '\tprimary = yes\n\n' +
                    '\tstrength_growth_per_century = 0.5\n' +
                    '\tmercenary_type = d_holy'+rcod+'_composition\n' +
                    '}\n')
        f.close()

    def parse_mercenaries_in_excel(self, ws, filename):
        f = open(filename, 'w')

        f.write("# Define mercenary types here now instead of in the static_composition.txt file\n" +
                "# Also remember to tell the landed title to use this mercenary type instead.\n" +
                "# Several titles can refer to the same type as well now.\n" +
                "###################################################\n" +
                "# Mercenary compositions\n" +
                "###################################################\n\n\n" +


                "### HOLY ORDERS AUTOMATICALLY GENERATED AND MAINTAINED BY THE SELIN PARSER v0.22 - Do not edit by hand!!! If a change is required, change the python code or the source Matrix !!! ###\n\n" +

                "# Total Troop Count (before levy_size)\n" +
                "# is based on the religion's Civilization and, if relevant, Ascendant\n" +

                "# Default levy size = 3 (base)\n\n" +

                "# Statist = +1\n" +
                "# Martial = +2\n" +
                "# Populist = -1\n" +
                "# Messianic = +1\n" +
                "# Traditional = -1\n" +
                "# Scholarly = -2\n\n" +

                "# Mainstream = +2\n" +
                "# Local = -2\n" +
                "# Heretical = 1\n\n")
        for religion_row in range(self._religion_row_begin, self._religion_row_end):
            self.write_mercenaries(f, ws, religion_row, self.col2num('AP'), self.col2num('AW'))
        f.close()

    def write_mercenaries(self, f, ws, religion_row, begin, end):
        multiplier = ws.cell(column=self.col2num('AP'), row=religion_row).value
        if multiplier is None or multiplier == 0:
            return  # we skip religion with no multiplier
        # else we create an entry
        rcod = ws.cell(column=self.col2num(self._religion_name_col), row=religion_row).value
        f.write("d_holy"+rcod+"_composition = {\n")
        for col in range(begin, end+1):
            header_value = ws.cell(column=col, row=5).value
            cell_value = ws.cell(column=col, row=religion_row).value
            if header_value is None or cell_value == 0:
                continue
            if isinstance(cell_value, float):
                f.write('\t' + header_value.encode('utf-8') + " = %0.2f\n" % cell_value)
            if isinstance(cell_value, int):
                f.write('\t' + header_value.encode('utf-8') + " = " + str(cell_value) + '\n')
        f.write('}\n')

    def col2num(self, col):
        num = 0
        for c in col:
            if c in string.ascii_letters:
                num = num * 26 + (ord(c.upper()) - ord('A')) + 1
        return num

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print "usage: python selin_parser.py filename.xlsx"
    else:
        selin_parser = selin_parser()
        workbook = load_workbook(filename=sys.argv[1], data_only=True)
        worksheet = workbook.get_sheet_by_name('Definition')
        selin_parser.parse_modifiers_in_excel(worksheet, 'modifiers.txt')
        selin_parser.parse_mercenary_titles_in_excel(worksheet, 'mercenary_titles.txt')
        selin_parser.parse_mercenaries_in_excel(worksheet, 'mercenaries.txt')